#!/usr/bin/env python3
"""
Generate PIU test workbooks into samples/generated/.

Run from repo root:
    python scripts/gen_test_workbooks.py

Produces:
    samples/generated/piu_full_build.xlsx       – every part, first valid values
    samples/generated/piu_location_sweep.xlsx   – parts with locations × all their location options
    samples/generated/mock_realistic_piu.xlsx   – hand-authored realistic PIU build
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dtm_buildsheet.paths import ensure_workspace
from dtm_buildsheet.config_store import load_config

# ── colours (match template_builder.py) ──────────────────────────────────────
_NAVY       = "1E2761"
_GOLD       = "C9A84C"
_LIGHT_GRAY = "F2F2F2"
_DARK_GRAY  = "595959"
_WHITE      = "FFFFFF"

# Column layout (same as template_builder.py):
# A(1) ✓  B(2) Part  C(3) New/Used  D(4) Source  E(5) Manufacturer
# F(6) Model/Part#  G(7) Location  H(8) Color  I(9) Qty  J(10) Lens  K(11) Notes
COL_HEADERS = [
    "✓", "Part", "New/Used", "Source", "Manufacturer",
    "Model / Part #", "Location", "Color", "Qty", "Lens", "Notes",
]
COL_WIDTHS = {1: 4, 2: 30, 3: 11, 4: 13, 5: 22, 6: 20, 7: 28, 8: 20, 9: 8, 10: 22, 11: 32}


# ── style helpers ─────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color=None, italic=False):
    return Font(bold=bold, size=size, color=color or "000000", italic=italic, name="Calibri")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(bottom=thin)


# ── header writer ─────────────────────────────────────────────────────────────

def _write_header(ws, build_info: dict) -> None:
    """Write rows 1–13: banner, info labels + values, vehicle detail section."""
    navy_fill  = _fill(_NAVY)
    gold_fill  = _fill(_GOLD)
    value_fill = _fill(_LIGHT_GRAY)
    label_font = _font(bold=True, size=9, color=_NAVY)

    # Row 1 — title
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "DTM FLEET — BUILD SHEET"
    c.font  = _font(bold=True, size=14, color=_WHITE)
    c.fill  = navy_fill
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — subtitle
    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "Complete all applicable fields. Leave unused parts blank."
    c.font  = _font(italic=True, size=9, color=_DARK_GRAY)
    c.fill  = value_fill
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 16

    def _info_row(row, left_label, left_value, right_label="", right_value=""):
        # Left label spans A:B (matches updated template_builder layout)
        lc = ws.cell(row=row, column=1, value=left_label)
        lc.font = label_font; lc.fill = gold_fill; lc.alignment = _align(h="right")
        ws.merge_cells(f"A{row}:B{row}")
        # Left value in C:F
        ws.merge_cells(f"C{row}:F{row}")
        vc = ws.cell(row=row, column=3, value=left_value or "")
        vc.fill = value_fill; vc.alignment = _align(h="left")
        # Right label in G, right value in H:K
        if right_label:
            rc = ws.cell(row=row, column=7, value=right_label)
            rc.font = label_font; rc.fill = gold_fill; rc.alignment = _align(h="right")
            ws.merge_cells(f"H{row}:K{row}")
            rvc = ws.cell(row=row, column=8, value=right_value or "")
            rvc.fill = value_fill; rvc.alignment = _align(h="left")
        ws.row_dimensions[row].height = 16

    _info_row(3, "AGENCY / DEPT:",    build_info.get("agency", ""),          "SALES REP:",         build_info.get("sales_rep", ""))
    _info_row(4, "PRIMARY CONTACT:",  build_info.get("contact", ""),         "QUOTE / ESTIMATE #:", build_info.get("quote", ""))
    _info_row(5, "PHONE:",            build_info.get("phone", ""),           "ADDITIONAL INFO:",   build_info.get("additional_info", ""))
    _info_row(6, "EMAIL:",            build_info.get("email", ""),           "VEHICLE TYPE:",      build_info.get("vehicle_type", "PIU"))

    # Row 7 — NEW / EXISTING banners
    for col, label, end_col in [(1, "NEW VEHICLE", 6), (7, "EXISTING VEHICLE", 11)]:
        c = ws.cell(row=7, column=col, value=label)
        c.font = _font(bold=True, size=9, color=_WHITE); c.fill = navy_fill
        c.alignment = _align(h="center")
        ws.merge_cells(f"{get_column_letter(col)}7:{get_column_letter(end_col)}7")
    ws.row_dimensions[7].height = 16

    # Rows 8–13 — vehicle detail fields (label A:B, value C:F, right label G, right value H:K)
    field_labels = ["UNIT ID:", "VIN:", "YEAR:", "MAKE:", "MODEL:", "SUB MODEL:"]
    field_keys   = ["UNIT ID",  "VIN",  "YEAR",  "MAKE",  "MODEL",  "SUB MODEL"]
    new_v      = build_info.get("new_vehicle", {})
    existing_v = build_info.get("existing_vehicle", {})
    for i, (label, key) in enumerate(zip(field_labels, field_keys), start=8):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = label_font; lc.fill = gold_fill; lc.alignment = _align(h="right")
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:F{i}")
        vc = ws.cell(row=i, column=3, value=new_v.get(key, ""))
        vc.fill = value_fill; vc.alignment = _align(h="left")
        rc = ws.cell(row=i, column=7, value=label)
        rc.font = label_font; rc.fill = gold_fill; rc.alignment = _align(h="right")
        ws.merge_cells(f"H{i}:K{i}")
        rvc = ws.cell(row=i, column=8, value=existing_v.get(key, ""))
        rvc.fill = value_fill; rvc.alignment = _align(h="left")
        ws.row_dimensions[i].height = 15


def _write_col_headers(ws, row: int) -> None:
    for col, label in enumerate(COL_HEADERS, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _font(bold=True, size=9, color=_WHITE)
        c.fill = _fill(_NAVY)
        c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 18


def _write_section_row(ws, row: int, label: str) -> None:
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font = _font(bold=True, size=9, color=_WHITE)
    c.fill = _fill(_NAVY)
    c.alignment = _align(h="left")
    ws.row_dimensions[row].height = 16


def _write_part_row(ws, row: int, part: dict, alt: bool = False) -> None:
    row_fill = _fill(_LIGHT_GRAY) if alt else None
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        if row_fill:
            cell.fill = row_fill
        cell.border = _thin_border()
        cell.alignment = _align(v="center")

    is_sub    = part.get("sub", False)
    part_font = _font(size=9, italic=is_sub, color=_DARK_GRAY if is_sub else "000000")
    ws.cell(row=row, column=2,  value=part["name"]).font = part_font
    ws.cell(row=row, column=3,  value=part.get("new_used", ""))
    ws.cell(row=row, column=4,  value=part.get("source", ""))
    ws.cell(row=row, column=5,  value=part.get("manufacturer", ""))
    ws.cell(row=row, column=6,  value=part.get("model", ""))
    ws.cell(row=row, column=7,  value=part.get("location", ""))
    ws.cell(row=row, column=8,  value=part.get("color", ""))
    ws.cell(row=row, column=9,  value=part.get("qty") or "")
    ws.cell(row=row, column=10, value=part.get("lens", ""))
    ws.cell(row=row, column=11, value=part.get("notes", ""))
    ws.row_dimensions[row].height = 15


# ── workbook factory ──────────────────────────────────────────────────────────

def _make_workbook(build_info: dict, parts_by_section: list[tuple[str, list[dict]]]) -> Workbook:
    """Create a complete, styled workbook from header info + section/part data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Build Sheet"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A15"

    _write_header(ws, build_info)
    _write_col_headers(ws, row=14)

    current_row = 15
    alt = False
    for section_label, parts in parts_by_section:
        _write_section_row(ws, current_row, section_label)
        current_row += 1
        for part in parts:
            _write_part_row(ws, current_row, part, alt=alt)
            current_row += 1
            alt = not alt

    return wb


# ── config loading ────────────────────────────────────────────────────────────

def _load_app_configs() -> dict:
    paths = ensure_workspace()
    wr = load_config("workbook_rules.json", paths)
    pl = load_config("parts_library.json", paths)

    mfg_map: dict[str, list] = {}
    mdl_map: dict[str, list] = {}
    for part in pl.get("parts", []):
        mfg   = (part.get("manufacturer") or "").strip()
        model = (part.get("model_number")  or "").strip()
        for ptype in part.get("compatible_types", []):
            if mfg and mfg not in mfg_map.get(ptype, []):
                mfg_map.setdefault(ptype, []).append(mfg)
            if model and model not in mdl_map.get(ptype, []):
                mdl_map.setdefault(ptype, []).append(model)

    return {
        "template_sections": wr.get("template_sections", []),
        "part_rules":        wr.get("part_rules", {}),
        "mfg_map":           mfg_map,
        "mdl_map":           mdl_map,
    }


def _first(lst, default=""):
    return lst[0] if lst else default


def _part_defaults(name: str, cfg: dict) -> dict:
    rule   = cfg["part_rules"].get(name, {})
    mfgs   = rule.get("manufacturer", []) or cfg["mfg_map"].get(name, [])
    models = rule.get("models", [])       or cfg["mdl_map"].get(name, [])
    return {
        "manufacturer": _first(mfgs),
        "model":        _first(models),
        "location":     _first(rule.get("locations", [])),
        "color":        _first(rule.get("colors", [])),
        "qty":          _first(rule.get("quantities", []), 1),
        "lens":         _first(rule.get("lens", [])),
    }


# ── build info dicts ──────────────────────────────────────────────────────────

_PIU_BASE = {
    "vehicle_type": "PIU",
    "new_vehicle": {
        "UNIT ID":   "TEST-001",
        "VIN":       "1FM5K8AR0NGA00001",
        "YEAR":      "2024",
        "MAKE":      "Ford",
        "MODEL":     "Police Interceptor Utility",
        "SUB MODEL": "AWD",
    },
    "existing_vehicle": {},
}

_FULL_BUILD_INFO = {
    **_PIU_BASE,
    "agency":          "DTM TEST — PIU FULL BUILD",
    "sales_rep":       "DTM",
    "contact":         "TEST AUTOMATION",
    "quote":           "TEST-PIU-FULL",
    "phone":           "",
    "email":           "",
    "additional_info": "Auto-generated — every part, first valid values.",
}

_LOCATION_SWEEP_INFO = {
    **_PIU_BASE,
    "agency":          "DTM TEST — PIU LOCATION SWEEP",
    "sales_rep":       "DTM",
    "contact":         "TEST AUTOMATION",
    "quote":           "TEST-PIU-LOCS",
    "phone":           "",
    "email":           "",
    "additional_info": "Auto-generated — parts × all their location options.",
}

_MOCK_REALISTIC_INFO = {
    "vehicle_type":    "PIU",
    "agency":          "Riverside County Sheriff",
    "sales_rep":       "J. Harmon",
    "contact":         "Sgt. M. Delgado — Fleet Mgmt",
    "quote":           "RCS-2026-041",
    "phone":           "(951) 555-0177",
    "email":           "mdelgado@riversidesheriff.org",
    "additional_info": "Rush order — unit needed by 06/15",
    "new_vehicle": {
        "UNIT ID":   "RCS-4217",
        "VIN":       "1FM5K8AR0NGA42170",
        "YEAR":      "2024",
        "MAKE":      "Ford",
        "MODEL":     "Police Interceptor Utility",
        "SUB MODEL": "4WD",
    },
    "existing_vehicle": {},
}

# Hand-authored realistic build.  Edit freely — only this dict is "data", the
# generator code above is the mechanism.
_MOCK_REALISTIC_PARTS: list[tuple[str, list[dict]]] = [
    ("PUSH BUMPER / FRONT STRUCTURE", [
        {"name": "Push Bumper",  "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA",   "model": "PB400-FPIUx",  "location": "",                      "color": "",         "qty": 1, "lens": "",              "notes": ""},
        {"name": "Pit Bar",      "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA",   "model": "PIT-COMP",     "location": "",                      "color": "",         "qty": 1, "lens": "",              "notes": ""},
        {"name": "Wing Wraps",   "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA",   "model": "WING-WRAP",    "location": "",                      "color": "",         "qty": 1, "lens": "",              "notes": ""},
    ]),
    ("FRONT WARNING LIGHTS", [
        {"name": "Forward Warning 1",  "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION",      "location": "GRILL",                 "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Forward Warning 2",  "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION",      "location": "CENTER PLATE OF PB",    "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Front Side Warning", "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION",      "location": "SIDE OF PUSH BUMPER",   "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Pit Bar Warning",    "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION",      "location": "PIT BARS",              "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Headlight Flasher",  "new_used": "NEW", "source": "DTM", "manufacturer": "SOUNDOFF", "model": "NFSHFC2L", "location": "",                   "color": "",         "qty": 1, "lens": "",              "notes": ""},
        {"name": "Front Scene",        "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION",      "location": "SPECIFY LOCATION",      "color": "White",    "qty": 2, "lens": "STANDARD LENSE", "notes": "Grille area"},
    ]),
    ("SIREN / SPEAKER / PA", [
        {"name": "Siren Speaker 1",        "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "SA315P",  "location": "TOP OF PUSH BUMPER",   "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Siren Speaker 2",        "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "SA315P",  "location": "UNDER PUSH BUMPER",    "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Siren Speaker Bracket",  "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA", "model": "SPK-BKT", "location": "FOLLOW SPEAKER LOCATION", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Pa Mic",                 "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "PA21",    "location": "",                     "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("FRONT SPECIALTY", [
        {"name": "Opticom",     "new_used": "NEW", "source": "DTM", "manufacturer": "GLOBAL TRAFFIC", "model": "790",  "location": "IN LIGHT BAR",  "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Cradle Point","new_used": "NEW", "source": "DTM", "manufacturer": "CRADLEPOINT",    "model": "R920", "location": "ON EQUIPMENT TRAY", "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("MIRROR & SIDE WARNING LIGHTS", [
        {"name": "Mirror Warning 1", "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "UNDER MIRROR",          "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Mirror Warning 2", "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "UNDER MIRROR",          "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Side Warning 1",   "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "REAR DOOR WINDOW",      "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Side Warning 2",   "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "UPPER CARGO WINDOW",    "color": "Red/Blue", "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Side Scene",       "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "ON RUNNING BOARD",      "color": "White",    "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
    ]),
    ("LIGHT BARS", [
        {"name": "Roof Light Bar",          "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN",   "model": "LEGACY",     "location": "",  "color": "Red/Blue",  "qty": 1, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Front Interior Light Bar","new_used": "NEW", "source": "DTM", "manufacturer": "SOUNDOFF", "model": "NFTVI2B",    "location": "",  "color": "Red/Blue",  "qty": 1, "lens": "",               "notes": ""},
    ]),
    ("REAR WARNING LIGHTS", [
        {"name": "Rear Warning 1",   "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "UPPER REAR WINDOW",      "color": "Red/Blue", "qty": 4, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Rear Warning 2",   "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "ON TAILGATE/LIFTGATE",   "color": "Red/Blue", "qty": 4, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Rear Scene",       "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION", "location": "ON BACK RACK",           "color": "White",    "qty": 2, "lens": "STANDARD LENSE", "notes": ""},
        {"name": "Tail Light Flasher","new_used": "NEW", "source": "DTM", "manufacturer": "SOUNDOFF","model": "NFSTLR", "location": "",                  "color": "",         "qty": 1, "lens": "",               "notes": ""},
    ]),
    ("RADAR & THERMAL", [
        {"name": "Radar",                   "new_used": "NEW", "source": "DTM", "manufacturer": "STALKER", "model": "DUAL DSR",   "location": "ON DASH",            "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Front Radar Antenna Mount","new_used": "NEW", "source": "DTM", "manufacturer": "STALKER", "model": "ANT BRKT",  "location": "DRIVER SIDE DASH",   "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Rear Radar Antenna Mount", "new_used": "NEW", "source": "DTM", "manufacturer": "STALKER", "model": "ANT BRKT",  "location": "DRIVER SIDE PILLAR", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Radar Interface Cable (To Camera)", "new_used": "NEW", "source": "DTM", "manufacturer": "STALKER", "model": "CAM CABLE", "location": "", "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("LIGHT CONTROLLER & ELECTRONICS", [
        {"name": "Light Controller", "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "CENCOM SAPPHIRE", "location": "ON EQUIPMENT TRAY", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Control Head",     "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "CAMCC",           "location": "IN CENTER CONSOLE",  "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Vehicle interface","new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "CAMCC-VI",        "location": "",                   "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("RADIO & COMMUNICATIONS", [
        {"name": "Radio 1",           "new_used": "NEW", "source": "AGENCY", "manufacturer": "MOTOROLA", "model": "APX8500",  "location": "ON EQUIPMENT TRAY",          "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Radio Mic Clip 1",  "new_used": "NEW", "source": "AGENCY", "manufacturer": "MOTOROLA", "model": "RMN5127",  "location": "ON DASH",                    "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Radio Antenna Cable","new_used": "NEW", "source": "DTM",   "manufacturer": "SPECIFY MFG","model": "COAX-15", "location": "REAR LEFT ROOF",            "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Radio Antenna Top",  "new_used": "NEW", "source": "DTM",   "manufacturer": "SPECIFY MFG","model": "SPECIFY MODEL", "location": "ON DASH",          "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("CENTER CONSOLE & COMPUTER", [
        {"name": "Center Console",  "new_used": "NEW", "source": "DTM", "manufacturer": "HAVIS",     "model": "C-VS-7000-FO",  "location": "",              "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Arm Rest",        "new_used": "NEW", "source": "DTM", "manufacturer": "HAVIS",     "model": "PKG-ARM-10",    "location": "",              "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Motion Attachment","new_used": "NEW", "source": "DTM", "manufacturer": "HAVIS",    "model": "PKG-MTG-1",     "location": "MOUNTED TO CONSOLE", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Docking Station", "new_used": "NEW", "source": "DTM", "manufacturer": "HAVIS",     "model": "DS-GTC-503",    "location": "SPECIFY DOCK",  "color": "", "qty": 1, "lens": "", "notes": "Panasonic Toughbook 55"},
    ]),
    ("PRINTER", [
        {"name": "Printer",       "new_used": "NEW", "source": "DTM", "manufacturer": "ZEBRA",       "model": "ZQ630+",     "location": "PRINTER ARMREST MOUNT", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Printer Mount", "new_used": "NEW", "source": "DTM", "manufacturer": "HAVIS",       "model": "PM-120",     "location": "",                      "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Printer Power", "new_used": "NEW", "source": "DTM", "manufacturer": "SPECIFY MFG", "model": "SPECIFY MODEL","location": "",                     "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("CAMERAS & DVR", [
        {"name": "Camera DVR",       "new_used": "NEW", "source": "DTM",    "manufacturer": "AXON", "model": "FLEET 3",        "location": "",                    "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Front Camera",     "new_used": "NEW", "source": "DTM",    "manufacturer": "AXON", "model": "FLEET 3 FRONT",  "location": "",                    "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Rear Camera",      "new_used": "NEW", "source": "DTM",    "manufacturer": "AXON", "model": "FLEET 3 REAR",   "location": "",                    "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Body Camera Dock", "new_used": "NEW", "source": "AGENCY", "manufacturer": "AXON", "model": "DOCK SOLO",      "location": "ON CENTER CONSOLE",   "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Rear Seat Camera", "new_used": "NEW", "source": "DTM",    "manufacturer": "AXON", "model": "FLEET INSIDE",   "location": "CENTER OF PARTITION", "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("PARTITIONS & PRISONER AREA", [
        {"name": "Front Partition",    "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA", "model": "PB400-FPIUx",  "location": "", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Replacement Rear Seat","new_used": "NEW", "source": "DTM", "manufacturer": "HAVIS","model": "DS-SEAT-FPI",  "location": "", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Rear Seat Lights 1", "new_used": "NEW", "source": "DTM", "manufacturer": "WHELEN", "model": "ION",         "location": "", "color": "White", "qty": 1, "lens": "", "notes": ""},
        {"name": "Floor Pan",          "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA", "model": "FP-PIU",      "location": "", "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("GUNLOCKS", [
        {"name": "Gunlock 1",         "new_used": "NEW", "source": "DTM", "manufacturer": "SANTA CRUZ", "model": "DUAL HANDCUFF",  "location": "OVERHEAD ON FRONT PARTITION", "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Gunlock Bracket 1", "new_used": "NEW", "source": "DTM", "manufacturer": "SANTA CRUZ", "model": "OHD BRACKET",    "location": "",                           "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Gunlock 2",         "new_used": "NEW", "source": "DTM", "manufacturer": "SETINA",     "model": "SINGLE SHOT GUN","location": "GUN LOCK POCKET",            "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("ELECTRICAL & POWER", [
        {"name": "Battery Tender", "new_used": "NEW", "source": "DTM", "manufacturer": "SPECIFY MFG", "model": "SPECIFY MODEL", "location": "UNDER HOOD", "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
    ("FINISHING & ACCESSORIES", [
        {"name": "Floor Mats",   "new_used": "NEW", "source": "AGENCY", "manufacturer": "WEATHERTECH",  "model": "PIU LINER SET",  "location": "FRONT AND REAR",       "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Seat Covers",  "new_used": "NEW", "source": "AGENCY", "manufacturer": "SPECIFY MFG",  "model": "SPECIFY MODEL",  "location": "DRIVER AND PASSENGER",  "color": "", "qty": 1, "lens": "", "notes": ""},
        {"name": "Decals",       "new_used": "NEW", "source": "AGENCY", "manufacturer": "SPECIFY MFG",  "model": "SPECIFY MODEL",  "location": "",                      "color": "", "qty": 1, "lens": "", "notes": "Per agency spec"},
        {"name": "Harness",      "new_used": "NEW", "source": "DTM",    "manufacturer": "SPECIFY MFG",  "model": "SPECIFY MODEL",  "location": "",                      "color": "", "qty": 1, "lens": "", "notes": ""},
    ]),
]


# ── generators ────────────────────────────────────────────────────────────────

def generate_piu_full_build(out_path: Path, cfg: dict) -> None:
    """All 130 parts, first valid values for every field."""
    parts_by_section: list[tuple[str, list[dict]]] = []
    for section in cfg["template_sections"]:
        label = section["label"]
        parts = []
        for entry in section.get("parts", []):
            name = entry.get("name", "")
            if not name:
                continue
            d = _part_defaults(name, cfg)
            parts.append({
                "name":         name,
                "sub":          entry.get("sub", False),
                "new_used":     "NEW",
                "source":       "DTM",
                "manufacturer": d["manufacturer"],
                "model":        d["model"],
                "location":     d["location"],
                "color":        d["color"],
                "qty":          d["qty"],
                "lens":         d["lens"],
                "notes":        "",
            })
        if parts:
            parts_by_section.append((label, parts))

    wb = _make_workbook(_FULL_BUILD_INFO, parts_by_section)
    wb.save(out_path)
    total = sum(len(p) for _, p in parts_by_section)
    print(f"  ✓ {out_path.name} — {total} parts across {len(parts_by_section)} sections")


def generate_piu_location_sweep(out_path: Path, cfg: dict) -> None:
    """For each part with location options: one row per location.
    Parts with no location options are omitted (they're covered by full_build)."""
    parts_by_section: list[tuple[str, list[dict]]] = []
    total = 0
    for section in cfg["template_sections"]:
        label = section["label"]
        parts = []
        for entry in section.get("parts", []):
            name = entry.get("name", "")
            if not name:
                continue
            rule  = cfg["part_rules"].get(name, {})
            locs  = rule.get("locations", [])
            if not locs:
                continue
            d = _part_defaults(name, cfg)
            for loc in locs:
                parts.append({
                    "name":         name,
                    "sub":          entry.get("sub", False),
                    "new_used":     "NEW",
                    "source":       "DTM",
                    "manufacturer": d["manufacturer"],
                    "model":        d["model"],
                    "location":     loc,
                    "color":        d["color"],
                    "qty":          d["qty"],
                    "lens":         d["lens"],
                    "notes":        "loc sweep",
                })
        if parts:
            parts_by_section.append((label, parts))
            total += len(parts)

    wb = _make_workbook(_LOCATION_SWEEP_INFO, parts_by_section)
    wb.save(out_path)
    print(f"  ✓ {out_path.name} — {total} rows (parts × location options)")


def generate_mock_realistic(out_path: Path) -> None:
    """Hand-authored realistic Riverside County Sheriff PIU build."""
    wb = _make_workbook(_MOCK_REALISTIC_INFO, _MOCK_REALISTIC_PARTS)
    wb.save(out_path)
    total = sum(len(p) for _, p in _MOCK_REALISTIC_PARTS)
    print(f"  ✓ {out_path.name} — {total} parts (hand-authored realistic build)")


# ── entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    out_dir = ROOT / "samples" / "generated"
    out_dir.mkdir(parents=True, exist_ok=True)

    print("Loading app configs …")
    cfg = _load_app_configs()

    print(f"\nGenerating workbooks → {out_dir.relative_to(ROOT)}/")
    generate_piu_full_build(    out_dir / "piu_full_build.xlsx",    cfg)
    generate_piu_location_sweep(out_dir / "piu_location_sweep.xlsx", cfg)
    generate_mock_realistic(    out_dir / "mock_realistic_piu.xlsx")
    print("\nDone.")


if __name__ == "__main__":
    main()
