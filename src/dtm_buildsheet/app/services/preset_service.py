from __future__ import annotations

import base64
import io
import json
import logging
import re
import uuid
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path

from ...domain.input_models import PartInput
from ...paths import AppPaths
from ...storage.local import LocalStorageProvider
from ...storage.safety import validate_safe_id
from ..adapters.wiring import delete_via_proposal, save_via_proposal

_log = logging.getLogger(__name__)


def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()

_PART_FIELDS = {
    "name", "include", "manufacturer", "part_number",
    "location", "raw_color", "quantity", "lens", "notes",
    "explicit_color_profile", "driver_color", "passenger_color", "center_color",
}

_PART_DEFAULTS: dict = {
    "include": True,
    "manufacturer": "",
    "part_number": "",
    "location": "",
    "raw_color": "",
    "quantity": 0,
    "lens": "",
    "notes": "",
    "explicit_color_profile": "",
    "driver_color": "",
    "passenger_color": "",
    "center_color": "",
}


def validate_preset_payload(payload: dict) -> dict:
    """Validate and normalise a raw preset dict.  Raises ValueError on invalid input."""
    if not isinstance(payload, dict):
        raise ValueError("Preset payload must be a JSON object")

    preset_id = str(payload.get("preset_id", "")).strip()
    validate_safe_id(preset_id, label="preset_id")

    label = str(payload.get("label", "")).strip()
    if not label:
        raise ValueError("Preset must have a non-empty 'label'")

    raw_parts = payload.get("parts", [])
    if not isinstance(raw_parts, list):
        raise ValueError("Preset 'parts' must be a list")

    normalised_parts = []
    for i, raw in enumerate(raw_parts):
        if not isinstance(raw, dict):
            raise ValueError(f"Preset parts[{i}] must be an object")
        name = str(raw.get("name", "")).strip()
        if not name:
            raise ValueError(f"Preset parts[{i}] must have a non-empty 'name'")
        part: dict = {"name": name}
        for field, default in _PART_DEFAULTS.items():
            part[field] = raw.get(field, default)
        normalised_parts.append(part)

    agency_ids = payload.get("agency_ids", [])
    if not isinstance(agency_ids, list):
        agency_ids = []
    build_types = payload.get("build_types", [])
    if not isinstance(build_types, list):
        build_types = []

    po = payload.get("placement_overrides")

    return {
        "schema_version": max(int(payload.get("schema_version", 1)), 2) if (agency_ids or build_types) else int(payload.get("schema_version", 1)),
        "preset_id": preset_id,
        "label": label,
        "description": str(payload.get("description", "")).strip(),
        "vehicle_types": [str(v) for v in payload.get("vehicle_types", [])],
        "agency_ids": [str(a) for a in agency_ids],
        "build_types": [str(b) for b in build_types],
        "tag": str(payload.get("tag", "")).strip(),
        "parts": normalised_parts,
        "placement_overrides": po if isinstance(po, dict) else {},
    }


def _load_dir(directory: Path, source: str) -> list[dict]:
    """Read all .json files from *directory* as preset summaries; skip corrupt files."""
    results = []
    if not directory.exists():
        return results
    for path in sorted(directory.glob("*.json")):
        # Skip dotfiles like .settings_etags.json — the shared-settings sync
        # writes its eTag cache into this directory, and it's not a preset.
        if path.name.startswith("."):
            continue
        try:
            raw = json.loads(path.read_text("utf-8"))
            validated = validate_preset_payload(raw)
            results.append({
                "preset_id": validated["preset_id"],
                "label": validated["label"],
                "description": validated["description"],
                "vehicle_types": validated["vehicle_types"],
                "agency_ids": validated.get("agency_ids", []),
                "build_types": validated.get("build_types", []),
                "tag": validated.get("tag", ""),
                "source": source,
            })
        except Exception:
            _log.warning("Skipping invalid preset file: %s", path, exc_info=True)
    return results


def list_presets(paths: AppPaths,
                 cloud_presets: list[dict] | None = None) -> list[dict]:
    """Return all available presets; workspace overrides bundled for the same preset_id.

    Priority (highest → lowest):
      1. workspace  — user-created; mutable; lives in user app-data dir
      2. cloud      — future community presets; pass as cloud_presets=[...]
                      when a cloud fetch is available (not yet implemented)
      3. bundled    — ships with the app; read-only in production

    In dev mode workspace_presets_dir == bundled_presets_dir (source tree), so all
    presets write directly to git and ship to end users as bundled presets.
    """
    by_id: dict[str, dict] = {}
    dev_mode = paths.workspace_presets_dir == paths.bundled_presets_dir

    if dev_mode:
        # Single directory: load once, label as workspace (all editable in dev)
        for item in _load_dir(paths.bundled_presets_dir, source="workspace"):
            by_id[item["preset_id"]] = item
    else:
        # Bundled presets first (lowest priority — overridden by anything above)
        for item in _load_dir(paths.bundled_presets_dir, source="bundled"):
            by_id[item["preset_id"]] = item

        # ── CLOUD HOOK ───────────────────────────────────────────────────────
        # When cloud/community presets are implemented, fetch them here and
        # pass the list in as cloud_presets.  Cloud overrides bundled but not
        # workspace, so add them after bundled and before workspace.
        for item in cloud_presets or []:
            item.setdefault("source", "cloud")
            by_id[item["preset_id"]] = item
        # ─────────────────────────────────────────────────────────────────────

        # Workspace presets win over everything
        for item in _load_dir(paths.workspace_presets_dir, source="workspace"):
            by_id[item["preset_id"]] = item

    return sorted(by_id.values(), key=lambda x: x["label"].lower())


def load_preset(preset_id: str, paths: AppPaths) -> list[PartInput]:
    """Load a preset by ID and return its parts as PartInput objects.

    Workspace overrides bundled when the same preset_id exists in both.
    Raises FileNotFoundError if the preset does not exist.
    Raises ValueError if the preset file is malformed.
    """
    validate_safe_id(preset_id, label="preset_id")

    for search_dir in (paths.workspace_presets_dir, paths.bundled_presets_dir):
        candidate = search_dir / f"{preset_id}.json"
        if candidate.exists():
            raw = json.loads(candidate.read_text("utf-8"))
            validated = validate_preset_payload(raw)
            return [_part_input_from_dict(p) for p in validated["parts"]]

    raise FileNotFoundError(f"Preset not found: {preset_id}")


def load_preset_dict(preset_id: str, paths: AppPaths) -> dict:
    """Load a preset by ID and return the full validated dict (including placement_overrides).

    Workspace overrides bundled when the same preset_id exists in both.
    Raises FileNotFoundError if the preset does not exist.
    Raises ValueError if the preset file is malformed.
    """
    validate_safe_id(preset_id, label="preset_id")

    for search_dir in (paths.workspace_presets_dir, paths.bundled_presets_dir):
        candidate = search_dir / f"{preset_id}.json"
        if candidate.exists():
            raw = json.loads(candidate.read_text("utf-8"))
            return validate_preset_payload(raw)

    raise FileNotFoundError(f"Preset not found: {preset_id}")


def _part_input_from_dict(d: dict) -> PartInput:
    return PartInput(
        name=d["name"],
        include=bool(d.get("include", True)),
        new_or_used="",
        source="",
        manufacturer=str(d.get("manufacturer", "")),
        part_number=str(d.get("part_number", "")),
        location=str(d.get("location", "")),
        raw_color=str(d.get("raw_color", "")),
        quantity=int(d.get("quantity", 0)),
        lens=str(d.get("lens", "")),
        notes=str(d.get("notes", "")),
        explicit_color_profile=str(d.get("explicit_color_profile", "")),
        driver_color=str(d.get("driver_color", "")),
        passenger_color=str(d.get("passenger_color", "")),
        center_color=str(d.get("center_color", "")),
    )


# ── helpers ───────────────────────────────────────────────────────────────────

def _is_bundled(preset_id: str, paths: AppPaths) -> bool:
    # In dev mode workspace IS the bundled dir — everything is editable.
    if paths.workspace_presets_dir == paths.bundled_presets_dir:
        return False
    return (paths.bundled_presets_dir / f"{preset_id}.json").exists()


def _workspace_path(preset_id: str, paths: AppPaths) -> Path:
    return paths.workspace_presets_dir / f"{preset_id}.json"


def _auto_name(payload: dict, paths: AppPaths) -> str:
    """Compute the auto-generated label from agency_ids, build_types, vehicle_types, tag."""
    from ..services.agency_service import load_agencies

    agency_ids = payload.get("agency_ids") or []
    build_types = payload.get("build_types") or []
    vehicle_types = payload.get("vehicle_types") or []
    tag = str(payload.get("tag", "")).strip()

    if agency_ids:
        agencies = load_agencies(paths)
        by_id = {a.agency_id: a.name for a in agencies}
        prefix = by_id.get(agency_ids[0], "Agency")
    else:
        prefix = "General"

    parts = [prefix]
    if len(build_types) == 1:
        parts.append(build_types[0])
    if vehicle_types:
        parts.append("/".join(vehicle_types))

    name = " ".join(parts)
    if not agency_ids and tag:
        name = f"{name} — {tag}"
    return name


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")[:40]


# ── write operations ──────────────────────────────────────────────────────────

def find_duplicate(payload: dict, paths: AppPaths, exclude_id: str = "") -> dict | None:
    """Return the first workspace preset that matches (agency_ids, build_types, vehicle_types, tag).

    Bundled presets are never considered duplicates — they cannot be overwritten.
    Returns None if no match is found.
    """
    candidate_agency   = sorted(str(a) for a in (payload.get("agency_ids") or []))
    candidate_builds   = sorted(str(b) for b in (payload.get("build_types") or []))
    candidate_vehicles = sorted(str(v) for v in (payload.get("vehicle_types") or []))
    candidate_tag      = str(payload.get("tag", "")).strip()

    for p in list_presets(paths):
        if p["preset_id"] == exclude_id:
            continue
        if p.get("source") != "workspace":
            continue  # never conflict with bundled/cloud presets
        if (sorted(p.get("agency_ids") or []) == candidate_agency and
                sorted(p.get("build_types") or []) == candidate_builds and
                sorted(p.get("vehicle_types") or []) == candidate_vehicles and
                str(p.get("tag", "")).strip() == candidate_tag):
            return p
    return None


def save_preset(payload: dict, paths: AppPaths, overwrite: bool = False) -> dict:
    """Create or update a preset in workspace/presets/. Auto-generates label.

    Enforces uniqueness by (agency_ids, build_types, vehicle_types, tag).
    Returns a conflict response when a duplicate exists unless *overwrite* is True
    or the payload already targets the duplicate's preset_id.
    """
    working = dict(payload)  # shallow copy — do not mutate caller's dict
    exclude_id = str(working.get("preset_id") or "").strip()

    dup = find_duplicate(working, paths, exclude_id=exclude_id)
    if dup and not overwrite:
        return {
            "ok": False,
            "conflict": True,
            "duplicate_preset_id": dup["preset_id"],
            "duplicate_label": dup["label"],
        }
    if dup and overwrite:
        working["preset_id"] = dup["preset_id"]

    if not working.get("preset_id"):
        base = _slug(_auto_name(working, paths) or "preset")
        working["preset_id"] = f"{base}_{uuid.uuid4().hex[:6]}"

    working["label"] = _auto_name(working, paths)

    try:
        validated = validate_preset_payload(working)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}

    preset_id = validated["preset_id"]
    serialized = json.dumps(validated, indent=2) + "\n"
    LocalStorageProvider().write_text(
        str(_workspace_path(preset_id, paths)),
        serialized,
    )
    proposal_result = save_via_proposal(
        target_file=f"presets/{preset_id}.json",
        serialized_content=serialized,
        summary=f"Update preset: {validated['label']}",
        category="general",
    )
    # Direct SP mirror — see agency_service comment.
    from .shared_work_service import save_setting_to_cloud_in_background
    save_setting_to_cloud_in_background(f"presets/{preset_id}.json", serialized)
    return {
        "ok": True,
        "preset_id": preset_id,
        "label": validated["label"],
        **proposal_result,
    }


def delete_preset(preset_id: str, paths: AppPaths) -> dict:
    """Delete a workspace preset. Returns error if it's a bundled preset.

    Propagates the deletion to cloud (and from there to other devices)
    via the proposal pipeline. Bundled presets never get a proposal —
    they're read-only and live in resources/presets/, not /Settings/.
    """
    try:
        validate_safe_id(preset_id, label="preset_id")
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    if _is_bundled(preset_id, paths):
        return {"ok": False, "error": "Cannot delete a bundled preset"}
    p = _workspace_path(preset_id, paths)
    if not p.exists():
        return {"ok": False, "error": "Preset not found"}
    # Capture the label before unlinking so the proposal summary is human-readable.
    label = preset_id
    try:
        label = (json.loads(p.read_text("utf-8")) or {}).get("label", preset_id) or preset_id
    except Exception:
        pass
    p.unlink()
    proposal_result = delete_via_proposal(
        target_file=f"presets/{preset_id}.json",
        summary=f"Delete preset: {label}",
        category="general",
    )
    # Belt-and-suspenders direct delete (see agency_service comment).
    from .shared_work_service import delete_setting_from_cloud
    delete_setting_from_cloud(f"presets/{preset_id}.json")
    return {"ok": True, **proposal_result}


def clone_preset(preset_id: str, paths: AppPaths) -> dict:
    """Copy a preset (workspace or bundled) with a new ID; strips agency_ids."""
    try:
        validate_safe_id(preset_id, label="preset_id")
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}

    original = None
    for search_dir in (paths.workspace_presets_dir, paths.bundled_presets_dir):
        candidate = search_dir / f"{preset_id}.json"
        if candidate.exists():
            original = json.loads(candidate.read_text("utf-8"))
            break
    if original is None:
        return {"ok": False, "error": "Preset not found"}

    new_id = f"clone_{uuid.uuid4().hex[:8]}"
    original["preset_id"] = new_id
    original["label"] = "Copy of " + original.get("label", "Preset")
    original["agency_ids"] = []  # strip agency assignment per plan

    try:
        validated = validate_preset_payload(original)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}

    LocalStorageProvider().write_text(
        str(_workspace_path(new_id, paths)),
        json.dumps(validated, indent=2) + "\n",
    )
    return {"ok": True, "preset_id": new_id, "label": validated["label"]}


def import_from_workbook(file_b64: str, paths: AppPaths) -> dict:
    """Decode base64 xlsx, parse parts via excel_reader, return preset payload."""
    import tempfile
    from ...inputs.excel_reader import load_input

    try:
        raw_bytes = base64.b64decode(file_b64)
    except Exception:
        return {"ok": False, "error": "Invalid base64 data"}

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(raw_bytes)
        tmp_path = Path(tmp.name)
    try:
        project = load_input(tmp_path)
    except Exception as exc:
        _log.warning("import_from_workbook: parse failed: %s", exc)
        return {"ok": False, "error": f"Could not parse workbook: {exc}"}
    finally:
        tmp_path.unlink(missing_ok=True)

    return {
        "ok": True,
        "parts": [asdict(p) for p in project.parts],
        "suggested_label": "Imported Preset",
        "vehicle_types": [],
        "build_types": [],
        "agency_ids": [],
    }


def export_preset_workbook(preset_id: str, paths: AppPaths) -> bytes:
    """Return an xlsx file populated with the preset's parts list."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    parts = load_preset(preset_id, paths)

    wb = Workbook()
    ws = wb.active
    ws.title = "Build Sheet"

    # Column layout matches template_builder so excel_reader can parse it back
    COL_HEADERS = ["✓", "Part", "New/Used", "Source", "Manufacturer",
                   "Model / Part #", "Location", "Color", "Qty", "Lens", "Notes"]
    COL_WIDTHS = {1: 4, 2: 30, 3: 11, 4: 13, 5: 22, 6: 20, 7: 28, 8: 20, 9: 8, 10: 22, 11: 32}

    navy = "1E2761"
    white = "FFFFFF"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write column headers at row 14 (matches template structure)
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(bold=True, size=9, color=white)
    for col, label in enumerate(COL_HEADERS, start=1):
        c = ws.cell(row=14, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[14].height = 18

    # Write part rows starting at row 15
    for row_idx, part in enumerate(parts, start=15):
        ws.cell(row=row_idx, column=1, value="✓" if part.include else "")
        ws.cell(row=row_idx, column=2, value=part.name)
        ws.cell(row=row_idx, column=3, value=part.new_or_used)
        ws.cell(row=row_idx, column=4, value=part.source)
        ws.cell(row=row_idx, column=5, value=part.manufacturer)
        ws.cell(row=row_idx, column=6, value=part.part_number)
        ws.cell(row=row_idx, column=7, value=part.location)
        ws.cell(row=row_idx, column=8, value=part.raw_color)
        ws.cell(row=row_idx, column=9, value=part.quantity or None)
        ws.cell(row=row_idx, column=10, value=part.lens)
        ws.cell(row=row_idx, column=11, value=part.notes)
        ws.row_dimensions[row_idx].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
