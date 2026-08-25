from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..domain import PartInput, ProjectInput
from ..domain.supply import normalized_supply_fields
from ..naming import canonical_name, safe_project_id


def _s(value) -> str:
    return "" if value is None else str(value).strip()


def _int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _bool_from_include(value) -> bool:
    if value is None:
        return True
    text = _s(value).lower()
    if not text:
        return True
    return text not in {"n", "no", "false", "0", "off"}


def _normalize_header(value: str) -> str:
    return _s(value).lower().replace("\n", " ").strip()


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        values = [_normalize_header(cell.value) for cell in row]
        if "part" not in values:
            continue
        index = {value: i for i, value in enumerate(values) if value}
        return row[0].row, index
    raise ValueError("Could not find parts header row")


def _col(row, header_map: dict[str, int], keys: list[str], fallback: int | None = None) -> str:
    for key in keys:
        idx = header_map.get(key)
        if idx is not None and idx < len(row):
            return _s(row[idx].value)
    if fallback is not None and fallback < len(row):
        return _s(row[fallback].value)
    return ""


def _parse_project_info(ws) -> dict:
    info: dict[str, object] = {}
    label_map = {
        "AGENCY / DEPT:": "Agency",
        "AGENCY/DEPT NAME:": "Agency",
        "PRIMARY CONTACT:": "PrimaryContact",
        "PHONE:": "Phone",
        "EMAIL:": "Email",
        "SALES REP:": "SalesRep",
        "QUOTE / ESTIMATE #:": "QuoteNumber",
        "QUOTE/ESTIMATE NUMBER:": "QuoteNumber",
        "ADDITIONAL INFO:": "AdditionalInfo",
        "BUILD TYPE:": "BuildType",
        "VEHICLE TYPE:": "BuildType",   # backward compat for old workbooks
    }
    vehicle_fields = {"UNIT ID:", "VIN:", "YEAR:", "MAKE:", "MODEL:", "SUB MODEL:"}

    new_vehicle: dict[str, str] = {}
    existing_vehicle: dict[str, str] = {}
    new_vehicle_col: int | None = None
    existing_vehicle_col: int | None = None

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=False):
        for cell in row:
            value = _s(cell.value)
            if value == "NEW VEHICLE":
                new_vehicle_col = cell.column
                continue
            if value == "EXISTING VEHICLE":
                existing_vehicle_col = cell.column
                continue
            if value in label_map:
                for next_cell in row[cell.column:]:
                    next_value = _s(next_cell.value)
                    if next_value and next_value != value:
                        info[label_map[value]] = next_value
                        break
                continue
            if value.rstrip(":") + ":" in vehicle_fields or value in vehicle_fields:
                field_name = value.rstrip(":")
                is_existing = (
                    existing_vehicle_col is not None
                    and cell.column >= existing_vehicle_col
                )
                for next_cell in row[cell.column:]:
                    next_value = _s(next_cell.value)
                    if next_value and next_value != value:
                        if is_existing:
                            existing_vehicle[field_name] = next_value
                        else:
                            new_vehicle[field_name] = next_value
                        break

    info["NewVehicle"] = new_vehicle
    info["ExistingVehicle"] = existing_vehicle

    model = new_vehicle.get("MODEL", existing_vehicle.get("MODEL", ""))
    model_upper = model.upper()
    if "PIU" in model_upper or "UTILITY" in model_upper:
        info["VehicleType"] = "PIU"
    elif "TAHOE" in model_upper:
        info["VehicleType"] = "TAHOE"
    elif "DURANGO" in model_upper:
        info["VehicleType"] = "DURANGO"
    elif "CHARGER" in model_upper:
        info["VehicleType"] = "CHARGER"
    else:
        info["VehicleType"] = model_upper.replace(" ", "_")[:24] or "UNKNOWN"

    raw_project_id = _s(info.get("QuoteNumber", "")) or _s(info.get("Agency", "Project"))
    info["ProjectID"] = safe_project_id(raw_project_id, fallback="PROJECT")
    return info


def load_input(path: Path) -> ProjectInput:
    wb = load_workbook(path, data_only=True)
    ws = wb["Build Sheet"]

    info = _parse_project_info(ws)
    header_row, headers = _find_header_row(ws)
    parts: list[PartInput] = []

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        name = canonical_name(_col(row, headers, ["part"], fallback=1))
        if not name or _normalize_header(name) == "part":
            continue

        new_or_used = _col(row, headers, ["new/used", "new / used"], fallback=2)
        source = _col(row, headers, ["source"], fallback=3)
        manufacturer = _col(row, headers, ["manufacturer"], fallback=4)
        part_number = _col(row, headers, ["model/part#", "model / part #", "part #"], fallback=5)
        location = canonical_name(_col(row, headers, ["location"], fallback=6))
        raw_color = _col(row, headers, ["color"], fallback=7)
        qty_raw = _col(row, headers, ["qty", "quantity"], fallback=8)
        lens = _col(row, headers, ["lens"], fallback=9)
        notes = _col(row, headers, ["notes"], fallback=10)
        explicit_color_profile = _col(row, headers, ["color profile", "light type", "color configuration"])
        driver_color = _col(row, headers, ["driver color", "driver-side color"])
        passenger_color = _col(row, headers, ["passenger color", "passenger-side color"])
        center_color = _col(row, headers, ["center color"])
        include = _bool_from_include(_col(row, headers, ["include", "selected", "✓"], fallback=0))

        if name == name.upper() and not any([new_or_used, manufacturer, part_number, location, raw_color]):
            continue

        has_data = any(
            [
                new_or_used,
                source,
                manufacturer,
                part_number,
                location,
                raw_color,
                qty_raw,
                lens,
                notes,
                explicit_color_profile,
                driver_color,
                passenger_color,
                center_color,
            ]
        )
        if not has_data:
            continue

        supply = normalized_supply_fields({"new_or_used": new_or_used, "source": source})
        parts.append(
            PartInput(
                name=name,
                include=include,
                new_or_used=supply["new_or_used"],
                source=supply["source"],
                supply_type=supply["supply_type"],
                customer_condition=supply["customer_condition"],
                customer_source=supply["customer_source"],
                manufacturer=manufacturer,
                part_number=part_number,
                location=location,
                raw_color=raw_color,
                quantity=_int(qty_raw),
                lens=lens,
                notes=notes,
                explicit_color_profile=explicit_color_profile,
                driver_color=driver_color,
                passenger_color=passenger_color,
                center_color=center_color,
            )
        )

    notes: dict[str, list[str]] = {}
    if "Notes" in wb.sheetnames:
        from ..ppt_helpers import NOTES_CATEGORIES
        _cat_set = {c.upper() for c in NOTES_CATEGORIES}
        current_cat: str | None = None
        ws_notes = wb["Notes"]
        for row in ws_notes.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            col_a = _s(row[0]).strip().upper() if row[0] else ""
            col_b = _s(row[1]).strip() if len(row) > 1 and row[1] else ""
            if col_a in _cat_set:
                current_cat = col_a
            elif col_b and current_cat:
                notes.setdefault(current_cat, []).append(col_b)

    return ProjectInput(info=info, parts=parts, notes=notes)
