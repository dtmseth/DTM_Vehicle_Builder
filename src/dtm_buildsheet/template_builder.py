"""Generates a blank build_sheet_template_v2.xlsx from current config data.

All structure and dropdown content is derived live from config files:
  workbook_rules.json  → template_sections + per-part workbook dropdowns
  parts_library.json   → parts-library compatibility data for the GUI
  vehicle_layouts.json → location dropdown options
  asset_manifest.json  → (future: size rules, etc.)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .config_store import load_config
from .paths import AppPaths, ensure_workspace

# ── brand colours ──────────────────────────────────────────────────────────────
_NAVY       = "1E2761"
_GOLD       = "C9A84C"
_LIGHT_GRAY = "F2F2F2"
_DARK_GRAY  = "595959"
_WHITE      = "FFFFFF"

# ── column layout ──────────────────────────────────────────────────────────────
# A  B     C         D       E               F             G          H      I    J     K
# ✓  Part  New/Used  Source  Manufacturer  Model/Part#  Location  Color  Qty  Lens  Notes
COL_WIDTHS  = {1:4, 2:30, 3:11, 4:13, 5:22, 6:20, 7:28, 8:20, 9:8, 10:22, 11:32}
COL_HEADERS = ["✓","Part","New/Used","Source","Manufacturer","Model / Part #",
               "Location","Color","Qty","Lens","Notes"]

FIELD_COL = {
    'new_used':     3,
    'manufacturer': 5,
    'models':       6,
    'location':     7,
    'colors':       8,
    'quantities':   9,
    'lens':         10,
}

NEW_USED = ["NEW","USED","REUSED","N/A"]


# ── config loaders ─────────────────────────────────────────────────────────────
def _load_configs(paths: AppPaths) -> dict:
    """Load and return all live config sources."""

    wr = load_config("workbook_rules.json", paths)
    pl = load_config("parts_library.json", paths)
    layouts = load_config("vehicle_layouts.json", paths)

    # template_sections: list of {label, parts:[{name, sub}]}
    template_sections = wr.get("template_sections", [])

    # part_rules: fallback per-part colors/qty/lens from original XML parse
    part_rules: dict = wr.get("part_rules", {})

    # Build manufacturer + model lookup from parts library for optional fallback.
    # part_type_name → {manufacturers: set, models: set}
    mfg_map: dict[str, set] = {}
    mdl_map: dict[str, set] = {}
    for part in pl.get("parts", []):
        mfg   = (part.get("manufacturer") or "").strip()
        model = (part.get("model_number")  or "").strip()
        for ptype in part.get("compatible_types", []):
            if mfg:
                mfg_map.setdefault(ptype, set()).add(mfg)
            if model:
                mdl_map.setdefault(ptype, set()).add(model)

    # Aggregate all unique location names from every vehicle / view
    all_locations: list[str] = sorted({
        loc_name
        for v in layouts.get("vehicles", {}).values()
        for view in v.get("views", {}).values()
        for loc_name in view.get("locations", {}).keys()
    })

    vehicle_types: list[str] = sorted(layouts.get("vehicles", {}).keys())

    return {
        "template_sections": template_sections,
        "part_rules":        part_rules,
        "mfg_map":           {k: sorted(v) for k, v in mfg_map.items()},
        "mdl_map":           {k: sorted(v) for k, v in mdl_map.items()},
        "all_locations":     all_locations,
        "vehicle_types":     vehicle_types,
    }


def _build_row_options(part_name: str, cfg: dict) -> dict:
    """Return per-field dropdown lists for a single part-type row."""
    rule = cfg["part_rules"].get(part_name, {})

    # Workbook rules are the authoritative dropdown source; parts library is only fallback.
    mfgs = rule.get("manufacturer", []) or cfg["mfg_map"].get(part_name) or []

    # Workbook rules are the authoritative dropdown source; parts library is only fallback.
    models = rule.get("models", []) or cfg["mdl_map"].get(part_name) or []

    # Locations: per-part-type from workbook_rules part_rules (XML-derived + GUI-editable)
    locs = rule.get("locations", [])

    # Colors / quantities / lens: from workbook_rules part_rules (no GUI editor yet)
    colors = rule.get("colors", [])
    qtys   = rule.get("quantities", [])
    lens   = rule.get("lens", [])

    return {
        "manufacturer": mfgs,
        "models":       models,
        "location":     locs,
        "colors":       colors,
        "quantities":   qtys,
        "lens":         lens,
    }


# ── style helpers ──────────────────────────────────────────────────────────────
def _font(bold=False, size=10, color=None, italic=False):
    return Font(bold=bold, size=size, color=color or "000000", italic=italic, name="Calibri")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(bottom=thin)


# ── data validation helpers ────────────────────────────────────────────────────
_MAX_INLINE = 240

def _inline_dv(options: list[str], allow_blank=True) -> DataValidation | None:
    if not options:
        return None
    joined = ",".join(str(o) for o in options)
    if len(joined) > _MAX_INLINE:
        joined = joined[:_MAX_INLINE].rsplit(",", 1)[0]
    return DataValidation(
        type="list",
        formula1=f'"{joined}"',
        allow_blank=allow_blank,
        showErrorMessage=False,
    )


# ── sheet builders ─────────────────────────────────────────────────────────────
def _write_header_section(ws, row_start: int, vehicle_types: list[str] | None = None) -> int:
    navy_fill  = _fill(_NAVY)
    gold_fill  = _fill(_GOLD)
    value_fill = _fill(_LIGHT_GRAY)
    label_font = _font(bold=True, size=9, color=_NAVY)

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "DTM FLEET — BUILD SHEET"
    c.font = _font(bold=True, size=14, color=_WHITE)
    c.fill = navy_fill
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "Complete all applicable fields. Leave unused parts blank."
    c.font = _font(italic=True, size=9, color=_DARK_GRAY)
    c.fill = _fill(_LIGHT_GRAY)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 16

    def _info_row(row, left_label, right_label=""):
        lc = ws.cell(row=row, column=1, value=left_label)
        lc.font = label_font; lc.fill = gold_fill; lc.alignment = _align(h="right")
        ws.merge_cells(f"B{row}:F{row}")
        ws.cell(row=row, column=2).fill = value_fill
        if right_label:
            rc = ws.cell(row=row, column=7, value=right_label)
            rc.font = label_font; rc.fill = gold_fill; rc.alignment = _align(h="right")
            ws.merge_cells(f"H{row}:K{row}")
            ws.cell(row=row, column=8).fill = value_fill
        ws.row_dimensions[row].height = 16

    _info_row(3, "AGENCY / DEPT:", "SALES REP:")
    _info_row(4, "PRIMARY CONTACT:", "QUOTE / ESTIMATE #:")
    _info_row(5, "PHONE:", "ADDITIONAL INFO:")
    _info_row(6, "EMAIL:", "VEHICLE TYPE:")

    # Vehicle type dropdown on H6
    if vehicle_types:
        vt_dv = _inline_dv(vehicle_types)
        if vt_dv:
            ws.add_data_validation(vt_dv)
            vt_dv.add(ws.cell(row=6, column=8))

    for col, label, end_col in [(1,"NEW VEHICLE",6),(7,"EXISTING VEHICLE",11)]:
        c = ws.cell(row=7, column=col, value=label)
        c.font = _font(bold=True, size=9, color=_WHITE); c.fill = navy_fill
        c.alignment = _align(h="center")
        ws.merge_cells(f"{get_column_letter(col)}7:{get_column_letter(end_col)}7")
    ws.row_dimensions[7].height = 16

    vehicle_fields = ["UNIT ID:","VIN:","YEAR:","MAKE:","MODEL:","SUB MODEL:"]
    for i, field in enumerate(vehicle_fields, start=8):
        for col_start, merge_range in [(1, f"B{i}:F{i}"), (7, f"H{i}:K{i}")]:
            c = ws.cell(row=i, column=col_start, value=field)
            c.font = label_font; c.fill = gold_fill; c.alignment = _align(h="right")
            ws.merge_cells(merge_range)
            ws.cell(row=i, column=col_start+1).fill = value_fill
        ws.row_dimensions[i].height = 15

    return 14


def _write_col_headers(ws, row: int):
    header_fill = _fill(_NAVY)
    header_font = _font(bold=True, size=9, color=_WHITE)
    for col, label in enumerate(COL_HEADERS, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = header_font; c.fill = header_fill
        c.alignment = _align(h="center")
    ws.row_dimensions[row].height = 18


def _write_part_rows(ws, row_start: int, template_sections: list) -> list[tuple[int, str]]:
    """Write section headers + part rows. Returns list of (excel_row, part_name)."""
    section_fill = _fill(_NAVY)
    section_font = _font(bold=True, size=9, color=_WHITE)
    alt_fill     = _fill(_LIGHT_GRAY)
    sub_font     = _font(size=9, italic=True, color=_DARK_GRAY)
    part_font    = _font(size=9)

    current_row = row_start
    alt = False
    data_rows: list[tuple[int, str]] = []

    for section in template_sections:
        label = section.get("label", "")
        parts = section.get("parts", [])

        ws.merge_cells(f"A{current_row}:K{current_row}")
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = section_font; c.fill = section_fill
        c.alignment = _align(h="left")
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for part_entry in parts:
            name   = part_entry.get("name", "")
            is_sub = part_entry.get("sub", False)
            row_fill = alt_fill if alt else None
            for col in range(1, 12):
                cell = ws.cell(row=current_row, column=col)
                if row_fill:
                    cell.fill = row_fill
                cell.border = _thin_border()
                cell.alignment = _align(v="center")

            part_cell = ws.cell(row=current_row, column=2, value=name)
            part_cell.font = sub_font if is_sub else part_font
            data_rows.append((current_row, name))
            ws.row_dimensions[current_row].height = 15
            current_row += 1
            alt = not alt

    return data_rows


def _apply_validations(ws, data_rows: list[tuple[int, str]], cfg: dict):
    """Apply per-row dropdown validations from live config data."""
    # Global New/Used
    nu_dv = _inline_dv(NEW_USED)
    if nu_dv:
        ws.add_data_validation(nu_dv)
        for excel_row, _ in data_rows:
            nu_dv.add(ws.cell(row=excel_row, column=FIELD_COL['new_used']))

    # Per-row: group identical option sets to reuse DV objects
    dv_cache: dict[tuple, DataValidation] = {}

    def get_dv(options: list[str]) -> DataValidation | None:
        key = tuple(options)
        if key not in dv_cache:
            dv = _inline_dv(options)
            if dv:
                ws.add_data_validation(dv)
                dv_cache[key] = dv
        return dv_cache.get(key)

    for excel_row, part_name in data_rows:
        opts = _build_row_options(part_name, cfg)

        for field, col_key in [
            ("manufacturer", "manufacturer"),
            ("models",       "models"),
            ("location",     "location"),
            ("colors",       "colors"),
            ("quantities",   "quantities"),
            ("lens",         "lens"),
        ]:
            options = opts.get(field, [])
            if options:
                dv = get_dv(options)
                if dv:
                    dv.add(ws.cell(row=excel_row, column=FIELD_COL[col_key]))


def _build_notes_sheet(wb: Workbook):
    ws = wb.create_sheet("Notes")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 80
    header_fill = _fill(_NAVY)
    for col, label in [(1, "#"), (2, "Note")]:
        c = ws.cell(row=1, column=col, value=label)
        c.font = _font(bold=True, color=_WHITE); c.fill = header_fill
        c.alignment = _align(h="center")
    for i in range(2, 22):
        ws.cell(row=i, column=1, value=i-1).alignment = _align(h="center")
        ws.cell(row=i, column=2).alignment = _align(wrap=True)
        ws.row_dimensions[i].height = 20


# ── public entry point ─────────────────────────────────────────────────────────
def build_template(paths: AppPaths | None = None, out_path: Path | None = None) -> Path:
    active_paths = paths or ensure_workspace()
    cfg = _load_configs(active_paths)

    if not cfg["template_sections"]:
        raise ValueError(
            "workbook_rules.json has no 'template_sections' key. "
            "Please regenerate from the GUI after saving part types."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Build Sheet"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A15"

    _write_header_section(ws, row_start=1, vehicle_types=cfg["vehicle_types"])
    _write_col_headers(ws, row=14)
    data_rows = _write_part_rows(ws, row_start=15, template_sections=cfg["template_sections"])
    _apply_validations(ws, data_rows, cfg)
    _build_notes_sheet(wb)

    if out_path is None:
        out_path = active_paths.workspace_dir / "build_sheet_template_v2.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
