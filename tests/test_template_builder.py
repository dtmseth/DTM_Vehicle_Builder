"""Baseline tests: template_builder generates a valid Excel workbook."""
from __future__ import annotations

import pytest
from openpyxl import load_workbook

from dtm_buildsheet.template_builder import build_template


@pytest.fixture(scope="module")
def template_path(tmp_path_factory, app_paths):
    out_dir = tmp_path_factory.mktemp("template")
    out_path = out_dir / "test_template.xlsx"
    return build_template(paths=app_paths, out_path=out_path)


def test_template_generates(template_path):
    assert template_path.exists(), f"Template file not created at {template_path}"


def test_template_has_xlsx_extension(template_path):
    assert template_path.suffix == ".xlsx"


def test_template_has_reasonable_size(template_path):
    size = template_path.stat().st_size
    assert size > 5_000, f"Template file suspiciously small: {size} bytes"


def test_template_is_openable_by_openpyxl(template_path):
    wb = load_workbook(str(template_path))
    assert wb is not None


def test_template_has_build_sheet(template_path):
    wb = load_workbook(str(template_path))
    assert "Build Sheet" in wb.sheetnames


def test_template_has_notes_sheet(template_path):
    wb = load_workbook(str(template_path))
    assert "Notes" in wb.sheetnames


def test_template_header_row_exists(template_path):
    wb = load_workbook(str(template_path))
    ws = wb["Build Sheet"]
    # Header row is row 14 per template_builder COL_HEADERS constant
    header_cells = [ws.cell(row=14, column=c).value for c in range(1, 12)]
    non_empty = [v for v in header_cells if v]
    assert len(non_empty) >= 5, f"Expected header row at row 14, got: {header_cells}"


def test_template_has_part_rows(template_path):
    wb = load_workbook(str(template_path))
    ws = wb["Build Sheet"]
    # Part rows start at row 15; at least a few should exist
    part_cells = [ws.cell(row=r, column=2).value for r in range(15, 60)]
    non_empty = [v for v in part_cells if v]
    assert len(non_empty) > 0, "Expected at least one part row in Build Sheet"


def test_template_has_data_validations(template_path):
    wb = load_workbook(str(template_path))
    ws = wb["Build Sheet"]
    assert len(ws.data_validations.dataValidation) > 0, "Expected dropdown validations in Build Sheet"


def test_template_is_idempotent(tmp_path_factory, app_paths):
    out1 = tmp_path_factory.mktemp("t1") / "template.xlsx"
    out2 = tmp_path_factory.mktemp("t2") / "template.xlsx"
    build_template(paths=app_paths, out_path=out1)
    build_template(paths=app_paths, out_path=out2)

    wb1 = load_workbook(str(out1))
    wb2 = load_workbook(str(out2))
    ws1 = wb1["Build Sheet"]
    ws2 = wb2["Build Sheet"]

    # Same sheet names
    assert set(wb1.sheetnames) == set(wb2.sheetnames)

    # Same header row
    h1 = [ws1.cell(row=14, column=c).value for c in range(1, 12)]
    h2 = [ws2.cell(row=14, column=c).value for c in range(1, 12)]
    assert h1 == h2
